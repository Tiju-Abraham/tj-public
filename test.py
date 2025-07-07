import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange

# --- CONFIGURATION ---

# Column mapping: (source_header, output_header)
column_mapping = [
    ("Full Name", "Name"),
    ("Age", "Age"),
    ("City Name", "City"),
    ("Country", "Country"),
    ("Splitpayment", ""),  # ignored
    ("SplitPercentage", ""),  # ignored
    ("", "add1"),  # synthetic
    ("TransactionID", "TransactionID"),  # row1-only
]

row1_only_headers = ["TransactionID"]

# Each group: (list of (source, output) pairs), comment text
row4_comment_groups = [
    ([ ("Full Name", "Name"), ("Age", "Age") ], "Personal Info"),
    ([ ("City Name", "City") ], "Location"),
    ([ ("Country", "Country"), ("", "add1") ], "Other Info"),
]

# --- HELPER FUNCTIONS ---
def normalize_header(header):
    return header.strip().lower() if header else ""

def get_source_header_index_map(ws):
    return {normalize_header(cell.value): idx for idx, cell in enumerate(ws[1], 1)}

def copy_cell_format(src_cell, dest_cell):
    dest_cell.font = src_cell.font.copy() if src_cell.font else Font()
    dest_cell.fill = src_cell.fill.copy() if src_cell.fill else PatternFill()
    dest_cell.border = src_cell.border.copy() if src_cell.border else Border()
    dest_cell.number_format = src_cell.number_format
    dest_cell.alignment = src_cell.alignment.copy() if src_cell.alignment else Alignment()

def build_output_columns(column_mapping, source_header_map, row1_only_headers):
    output_columns = []
    for src, out in column_mapping:
        is_row1_only = out in row1_only_headers
        src_norm = normalize_header(src)
        src_idx = source_header_map.get(src_norm) if src else None
        output_columns.append({
            "source": src,
            "output": out,
            "is_row1_only": is_row1_only,
            "source_idx": src_idx,
        })
    return output_columns

def write_headers(ws, output_columns):
    for col_idx, col in enumerate(output_columns, 1):
        ws.cell(row=1, column=col_idx, value=col["output"] or col["source"])

def write_true_false_row(ws, output_columns):
    for col_idx, col in enumerate(output_columns, 1):
        match = normalize_header(col["source"]) == normalize_header(col["output"])
        ws.cell(row=2, column=col_idx, value="TRUE" if match else "FALSE")

def write_source_headers(ws, output_columns):
    for col_idx, col in enumerate(output_columns, 1):
        if not col["is_row1_only"]:
            ws.cell(row=3, column=col_idx, value=col["source"])

def write_row4_comments(ws, output_columns, row4_comment_groups):
    col_map = [(col["source"], col["output"]) for col in output_columns]
    col_idx = 1 
    for group, comment in row4_comment_groups:
        # Find start and end col for this group
        group_indices = []
        for i, (src, out) in enumerate(col_map):
            for gsrc, gout in group:
                if normalize_header(src) == normalize_header(gsrc) and normalize_header(out) == normalize_header(gout):
                    group_indices.append(i+1)
        if not group_indices:
            continue
        start = min(group_indices)
        end = max(group_indices)
        ws.merge_cells(start_row=4, start_column=start, end_row=4, end_column=end)
        cell = ws.cell(row=4, column=start)
        cell.value = comment
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Remove borders inside merged region, keep only on start/end
        thin = Side(border_style="thin", color="000000")
        for c in range(start, end+1):
            cell = ws.cell(row=4, column=c)
            if c == start:
                cell.border = Border(left=thin, top=thin, bottom=thin, right=Border().right)
            elif c == end:
                cell.border = Border(right=thin, top=thin, bottom=thin, left=Border().left)
            else:
                cell.border = Border(top=Border().top, bottom=Border().bottom)

def copy_data(source_ws, dest_ws, output_columns, start_row_dest=5, row1_only_headers=None):
    if row1_only_headers is None:
        row1_only_headers = []
    for row_idx, src_row in enumerate(source_ws.iter_rows(min_row=2), start_row_dest):
        for col_idx, col in enumerate(output_columns, 1):
            if col["is_row1_only"]:
                continue  # skip row1-only columns in data
            if col["source_idx"]:
                src_cell = src_row[col["source_idx"] - 1]
                dest_cell = dest_ws.cell(row=row_idx, column=col_idx, value=src_cell.value)
                copy_cell_format(src_cell, dest_cell)
            else:
                dest_ws.cell(row=row_idx, column=col_idx, value=None)

def autofit_columns(ws, output_columns, max_row):
    for col_idx, col in enumerate(output_columns, 1):
        max_length = 0
        for row in range(1, max_row+1):
            val = ws.cell(row=row, column=col_idx).value
            if val:
                max_length = max(max_length, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

def main():
    # Load source workbook
    source_wb = openpyxl.load_workbook("source.xlsx")
    source_ws = source_wb.active

    # Normalize and map source headers
    source_header_map = get_source_header_index_map(source_ws)

    # Build output columns
    output_columns = build_output_columns(column_mapping, source_header_map, row1_only_headers)

    # Create destination workbook
    dest_wb = openpyxl.Workbook()
    dest_ws = dest_wb.active

    # Write headers and rows
    write_headers(dest_ws, output_columns)
    write_true_false_row(dest_ws, output_columns)
    write_source_headers(dest_ws, output_columns)
    write_row4_comments(dest_ws, output_columns, row4_comment_groups)
    copy_data(source_ws, dest_ws, output_columns, start_row_dest=5, row1_only_headers=row1_only_headers)
    autofit_columns(dest_ws, output_columns, max_row=5+source_ws.max_row-1)

    # Save destination workbook
    dest_wb.save("destination.xlsx")
    print("destination.xlsx created successfully.")

if __name__ == "__main__":
    main() 